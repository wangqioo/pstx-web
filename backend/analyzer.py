# -*- coding: utf-8 -*-
"""
PSTX 原理图分析工具 v1.2
解析 Cadence Packager-XL 导出的 pstxprt.dat / pstxnet.dat

功能：BOM 管理 / 网络拓扑 / DRC / 电容降额 / 电阻检查 / 元件查询 / Excel 导出

依赖：pip install openpyxl
运行：python pstx_analyzer.py
"""

import os
import re
from collections import Counter, defaultdict
from pathlib import Path
from typing import Dict, List, Optional, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


# ══════════════════════════════════════════════════════════
# 零、页码解析辅助（内联自 pstx_page_logic.py）
# ══════════════════════════════════════════════════════════

_PAGE_TOKEN_RE = re.compile(
    r'(?<![A-Z0-9])PAGE(?:[_\-/ ]*)(\d+)([A-Z]?)(?![A-Z0-9])',
    re.IGNORECASE,
)
_PATH_SEGMENT_RE = re.compile(
    r'^(?P<head>.+?)\((?P<view>[^)]+)\)\s*:\s*(?P<tail>.+)$',
    re.IGNORECASE,
)
_SECTION_PATH_RE = re.compile(
    r'(?ims)^\s*SECTION_NUMBER\s+(?P<num>\d+)\s*\n\s*\'(?P<path>[^\']+)\'\s*:',
)
_PAGE_NUMBER_LINE_RE = re.compile(
    r"""^\s*["']?PAGE_NUMBER["']?\s*(?:=|:)\s*["']?(?P<value>[A-Z0-9_./ -]+?)["']?\s*[;,]?\s*$""",
    re.IGNORECASE,
)


def _natural_sort_key(value: str):
    parts = re.split(r'(\d+)', str(value or '').upper())
    return [int(p) if p.isdigit() else p for p in parts]


def _normalize_page_token(match: re.Match) -> str:
    num = str(int(match.group(1)))
    suffix = match.group(2).upper()
    return f'PAGE{num}{suffix}'


def _normalize_page_label(page_label: str) -> str:
    value = str(page_label or '').strip().upper()
    if not value:
        return ''
    matches = list(_PAGE_TOKEN_RE.finditer(value))
    if not matches:
        return value
    normalized = [_normalize_page_token(m) for m in matches]
    return normalized[0] if len(normalized) == 1 else ' / '.join(normalized)


def _coerce_page_number(value: str) -> str:
    text = str(value or '').strip()
    if not text:
        return ''
    if not text.upper().startswith('PAGE'):
        text = f'PAGE{text}'
    return _normalize_page_label(text)


def _clean_page_csv_value(value: str) -> str:
    text = str(value or '').strip().rstrip(';,').strip()
    if len(text) >= 2 and text[0] == text[-1] and text[0] in {'"', "'"}:
        text = text[1:-1].strip()
    return text


def _iter_text_with_fallback_encodings(file_path) -> List[str]:
    """支持 utf-16 等多种编码读取文件，依次尝试，返回去重后的文本列表"""
    try:
        raw_bytes = Path(file_path).read_bytes()
    except OSError:
        return []
    texts: List[str] = []
    seen: set = set()
    for enc in ['utf-8-sig', 'utf-16', 'utf-16-le', 'utf-16-be', 'utf-8', 'gb18030', 'cp936']:
        try:
            text = raw_bytes.decode(enc)
        except UnicodeDecodeError:
            continue
        if text and text not in seen:
            seen.add(text)
            texts.append(text)
    fallback = raw_bytes.decode('utf-8', errors='replace')
    if fallback and fallback not in seen:
        texts.append(fallback)
    return texts


def _extract_page_number_from_text(text: str) -> str:
    if not text:
        return ''
    for raw_line in text.splitlines():
        line = raw_line.strip()
        if not line:
            continue
        match = _PAGE_NUMBER_LINE_RE.match(line)
        if match:
            pn = _coerce_page_number(_clean_page_csv_value(match.group('value')))
            if pn:
                return pn
    rows = []
    for raw_line in text.splitlines():
        parts = [_clean_page_csv_value(p) for p in raw_line.split(',')]
        rows.append(parts)
        for idx, part in enumerate(parts):
            if part.upper() != 'PAGE_NUMBER':
                continue
            for follower in parts[idx + 1:]:
                pn = _coerce_page_number(_clean_page_csv_value(follower))
                if pn:
                    return pn
    for row_idx, parts in enumerate(rows):
        header_idxs = [i for i, p in enumerate(parts) if p.upper() == 'PAGE_NUMBER']
        for col_idx in header_idxs:
            for data_row in rows[row_idx + 1:]:
                if col_idx >= len(data_row):
                    continue
                pn = _coerce_page_number(_clean_page_csv_value(data_row[col_idx]))
                if pn:
                    return pn
    for regex in [
        re.compile(r'(?im)["\']?PAGE_NUMBER["\']?\s*[,=:\t;]\s*["\']?([A-Z0-9_./ -]+?)["\']?\s*[;,]?(?:$|\r|\n)'),
        re.compile(r'(?im)^["\']?PAGE_NUMBER["\']?\s*[,;\t]\s*["\']?([A-Z0-9_./ -]+?)["\']?\s*[;,]?(?:$|\r|\n)'),
    ]:
        m = regex.search(text)
        if m:
            pn = _coerce_page_number(_clean_page_csv_value(m.group(1)))
            if pn:
                return pn
    return ''


def _read_page_number_from_csv(csv_path) -> str:
    for text in _iter_text_with_fallback_encodings(csv_path):
        pn = _extract_page_number_from_text(text)
        if pn:
            return pn
    return ''


def _iter_page_csv_paths(project_root: Path) -> List[Path]:
    candidates: Dict[str, Path] = {}
    direct_sch = project_root / 'sch_1'
    if direct_sch.is_dir():
        for csv_path in direct_sch.iterdir():
            if (csv_path.is_file() and csv_path.suffix.lower() == '.csv'
                    and csv_path.stem.lower().startswith('page')):
                candidates[str(csv_path.resolve())] = csv_path
    for csv_path in project_root.rglob('page*.csv'):
        if csv_path.is_file() and csv_path.parent.name.lower() == 'sch_1':
            candidates[str(csv_path.resolve())] = csv_path
    return sorted(candidates.values(), key=lambda p: _natural_sort_key(str(p)))


def _build_page_csv_index(project_root: str) -> Dict:
    root = Path(project_root).expanduser()
    index = {
        'root': str(root), 'by_logical_page': defaultdict(list),
        'warnings': [], 'count': 0, 'scanned': 0, 'skipped_paths': [],
    }
    if not project_root or not root.exists():
        if project_root:
            index['warnings'].append(f'项目根路径不存在：{root}')
        return index
    csv_paths = _iter_page_csv_paths(root)
    index['scanned'] = len(csv_paths)
    for csv_path in csv_paths:
        real_page = _coerce_page_number(csv_path.stem)
        if not real_page:
            index['skipped_paths'].append(str(csv_path))
            continue
        logical_page = _read_page_number_from_csv(csv_path)
        if not logical_page:
            index['skipped_paths'].append(str(csv_path))
            continue
        index['by_logical_page'][logical_page].append({
            'path': str(csv_path), 'resolved_page': real_page,
            'is_root_sch1': csv_path.parent == (root / 'sch_1'),
        })
        index['count'] += 1
    if index['scanned'] == 0:
        index['warnings'].append(f'未在项目根路径下找到任何 sch_1/page*.csv：{root}')
    elif index['count'] == 0:
        samples = '；'.join(index['skipped_paths'][:3])
        index['warnings'].append(f'已扫描 {index["scanned"]} 个 page*.csv，但没有读出任何 PAGE_NUMBER' +
                                  (f'；例如：{samples}' if samples else ''))
    return index


def _parse_page_map_line(raw_line: str) -> Optional[Dict]:
    parts = re.split(r'\s+', str(raw_line or '').strip(), maxsplit=2)
    if len(parts) < 3:
        return None
    lp = _coerce_page_number(parts[0])
    rp = _coerce_page_number(parts[1])
    if not lp or not rp:
        return None
    return {'logical_page': lp, 'real_page': rp, 'page_name': parts[2].strip()}


def _build_page_map_index(project_root: str) -> Dict:
    root = Path(project_root).expanduser()
    index = {'root': str(root), 'by_logical_page': defaultdict(list), 'warnings': [], 'count': 0}
    if not project_root or not root.exists():
        return index
    file_paths = []
    direct = root / 'sch_1' / 'page.map'
    if direct.is_file():
        file_paths.append(direct)
    for path in root.rglob('page.map'):
        if path.is_file() and path not in file_paths:
            file_paths.append(path)
    for path in file_paths:
        matched = False
        for text in _iter_text_with_fallback_encodings(path):
            for raw_line in text.splitlines():
                parsed = _parse_page_map_line(raw_line)
                if not parsed:
                    continue
                lp = parsed['logical_page']
                rp = parsed['real_page']
                index['by_logical_page'][lp].append({
                    'path': str(path), 'logical_page': lp, 'resolved_page': rp,
                    'page_name': parsed['page_name'],
                    'is_root_sch1': path.parent == (root / 'sch_1'),
                })
                index['count'] += 1
                matched = True
            if matched:
                break
    return index


def _resolve_unique_real_page(index: Optional[Dict], logical_page: str) -> Tuple[str, str]:
    if not index or not logical_page:
        return '', 'none'
    entries = index.get('by_logical_page', {}).get(logical_page, [])
    if not entries:
        return '', 'none'
    real_pages = sorted({e.get('resolved_page', '') for e in entries if e.get('resolved_page')},
                        key=_natural_sort_key)
    if len(real_pages) != 1:
        return '', 'ambiguous'
    return real_pages[0], 'unique'


def _extract_path_segments(path_text: str) -> List[Dict]:
    raw = str(path_text or '').strip()
    if not raw:
        return []
    segments = []
    for chunk in [s.strip() for s in raw.split('@') if s.strip()]:
        match = _PATH_SEGMENT_RE.match(chunk)
        if not match:
            continue
        head = match.group('head').strip()
        view = match.group('view').strip()
        tail = match.group('tail').strip()
        pm = _PAGE_TOKEN_RE.search(tail)
        if not pm:
            continue
        lib, _, cell = head.rpartition('.')
        segments.append({
            'raw': chunk, 'head': head, 'lib': lib.strip(),
            'cell': (cell or head).strip(), 'view': view,
            'raw_page': _normalize_page_token(pm), 'tail': tail,
        })
    return segments


def _extract_section_paths(block_text: str) -> List[Dict]:
    entries = []
    for m in _SECTION_PATH_RE.finditer(str(block_text or '')):
        entries.append({'section_number': m.group('num'), 'path': m.group('path').strip()})
    return entries


def _select_component_page_source(block_text: str, attrs: Dict) -> Tuple[str, str]:
    section_paths = _extract_section_paths(block_text)
    if section_paths:
        preferred = next((e for e in section_paths if e.get('section_number') == '1'), section_paths[0])
        path_text = preferred.get('path', '').strip()
        if path_text:
            return path_text, 'section_path'
    c_path = str(attrs.get('C_PATH', '')).strip()
    if c_path:
        return c_path, 'c_path'
    drawing = str(attrs.get('DRAWING', '')).strip()
    if drawing:
        return drawing, 'drawing'
    return '', 'none'


def _extract_top_level_logical_page(path_text: str) -> str:
    segments = _extract_path_segments(path_text)
    for seg in segments:
        if seg.get('view', '').upper() == 'SCH_1':
            return seg.get('raw_page', '')
    if segments:
        return segments[0].get('raw_page', '')
    return _normalize_page_label(path_text).split(' / ')[0] if path_text else ''


def _extract_submodule_page(path_text: str) -> str:
    sch_segs = [s for s in _extract_path_segments(path_text) if s.get('view', '').upper() == 'SCH_1']
    return sch_segs[1].get('raw_page', '') if len(sch_segs) == 2 else ''


def _pick_top_schematic_segment(path_text: str, page_map_index: Optional[Dict],
                                 page_csv_index: Optional[Dict]) -> Dict:
    sch_segs = [s for s in _extract_path_segments(path_text) if s.get('view', '').upper() == 'SCH_1']
    if not sch_segs:
        return {}
    # 优先找与项目根目录同名的模块
    root_name = ''
    for idx in [page_map_index, page_csv_index]:
        if idx and idx.get('root'):
            try:
                root_name = Path(str(idx['root'])).name.upper()
                break
            except Exception:
                pass
    if root_name:
        exact = [s for s in sch_segs if s.get('cell', '').upper() == root_name]
        if exact:
            return exact[0]
    # 优先找在 root sch_1 中有页码的
    root_pages = set()
    for idx in [page_map_index, page_csv_index]:
        if idx:
            for lp, entries in idx.get('by_logical_page', {}).items():
                if any(e.get('is_root_sch1') for e in entries):
                    root_pages.add(lp)
    if root_pages:
        root_matches = [s for s in sch_segs if s.get('raw_page', '') in root_pages]
        if root_matches:
            return root_matches[0]
    return sch_segs[0]


def _resolve_component_page(comp: Dict, page_map_index: Optional[Dict],
                             page_csv_index: Optional[Dict]) -> str:
    logical_path = str(comp.get('page_path_raw', '') or comp.get('drawing', ''))
    top_seg = _pick_top_schematic_segment(logical_path, page_map_index, page_csv_index)
    top_logical = top_seg.get('raw_page', '') or _extract_top_level_logical_page(logical_path)
    if not top_logical:
        return ''
    pm_real, _ = _resolve_unique_real_page(page_map_index, top_logical)
    csv_real, _ = _resolve_unique_real_page(page_csv_index, top_logical)
    return pm_real or csv_real or top_logical


def resolve_component_pages(components: Dict, project_root: str = '') -> List[str]:
    """用 page.map / page*.csv 把逻辑页转换为真实页，返回警告列表"""
    if not project_root:
        # 无项目根目录：PHYS_PAGE 已在解析时写入，这里只补全缺失项
        for comp in components.values():
            if not comp.get('page'):
                lp = _extract_top_level_logical_page(
                    str(comp.get('page_path_raw', '') or comp.get('drawing', '')))
                comp['page'] = lp
                comp['page_logical'] = lp
        return []
    pm_index = _build_page_map_index(project_root)
    csv_index = _build_page_csv_index(project_root)
    warnings = list(pm_index.get('warnings', [])) + list(csv_index.get('warnings', []))
    for comp in components.values():
        # 若已有 PHYS_PAGE，直接用，不再覆盖
        if comp.get('page_real'):
            continue
        logical_path = str(comp.get('page_path_raw', '') or comp.get('drawing', ''))
        top_seg = _pick_top_schematic_segment(logical_path, pm_index, csv_index)
        top_logical = top_seg.get('raw_page', '') or _extract_top_level_logical_page(logical_path)
        pm_real, _ = _resolve_unique_real_page(pm_index, top_logical)
        csv_real, _ = _resolve_unique_real_page(csv_index, top_logical)
        real_page = pm_real or csv_real or ''
        comp['page'] = real_page or top_logical
        comp['page_logical'] = top_logical
        comp['page_real'] = real_page
    return warnings




# ══════════════════════════════════════════════════════════
# 一、PST 文件解析
# ══════════════════════════════════════════════════════════

def _join_continuations(text: str) -> str:
    normalized = str(text or '').replace('\r\n', '\n').replace('\r', '\n')
    lines = normalized.split('\n')
    result, buf = [], ''
    for line in lines:
        stripped = line.rstrip()
        if stripped.endswith('~'):
            buf += stripped[:-1]
        else:
            buf += line
            result.append(buf)
            buf = ''
    if buf:
        result.append(buf)
    return '\n'.join(result)


def _extract_attrs(text: str) -> Dict[str, str]:
    attrs = {}
    for m in re.finditer(r"\b([A-Z][A-Z0-9_]*)\s*=\s*'([^']*)'", text):
        key, val = m.group(1), m.group(2)
        if key not in attrs:
            attrs[key] = val
    return attrs


def _get_comp_type(refdes: str, part_name: str) -> str:
    pn = part_name.lower()
    type_rules = [
        (['cap_pol'],                           'CAP_POL'),
        (['cap_hdl', 'cap_'],                   'CAP'),
        (['res_hdl', 'res_'],                   'RES'),
        (['ind_hdl', 'ind_', 'ferrite', 'fer_hdl', 'fb_hdl'], 'IND'),
        (['osc_', 'crystal', 'xtal'],           'XTAL'),
        (['conn_', 'connector'],                'CONN'),
        (['led_'],                              'LED'),
        (['diode', '_d_hdl'],                   'DIODE'),
        (['mosfet', 'mos_', 'nmos', 'pmos', 'nfet', 'pfet'], 'FET'),
        (['bjt', 'transistor', 'npn', 'pnp'],  'BJT'),
        (['fuse'],                              'FUSE'),
        (['sw_hdl', 'switch'],                  'SWITCH'),
        (['testpoint', 'test_point', 'tp_hdl'], 'TESTPOINT'),
        (['transformer', 'xfmr'],              'TRANSFORMER'),
    ]
    for keywords, ctype in type_rules:
        if any(k in pn for k in keywords):
            return ctype
    prefix = (re.match(r'[A-Za-z]+', refdes) or re.match(r'', '')).group(0).upper()
    prefix_map = {
        'C': 'CAP', 'PC': 'CAP', 'R': 'RES', 'L': 'IND', 'FB': 'IND',
        'U': 'IC', 'J': 'CONN', 'P': 'CONN', 'CN': 'CONN', 'Q': 'FET',
        'D': 'DIODE', 'LED': 'LED', 'Y': 'XTAL', 'F': 'FUSE',
        'SW': 'SWITCH', 'TP': 'TESTPOINT', 'T': 'TRANSFORMER',
    }
    return prefix_map.get(prefix, 'IC')


def parse_pstxprt(content: str) -> Dict[str, dict]:
    text = _join_continuations(content)
    components = {}
    for block in re.split(r'\nPART_NAME\n', text)[1:]:
        m = re.match(r"(\S+)\s+'([^']+)'", block.split('\n')[0].strip())
        if not m:
            continue
        refdes, part_name = m.group(1), m.group(2)
        attrs = _extract_attrs(block)
        page_path_raw, page_path_source = _select_component_page_source(block, attrs)
        logical_page = _extract_top_level_logical_page(page_path_raw or attrs.get('DRAWING', ''))
        # PHYS_PAGE 是工程师印刷原理图上看到的实际页码。
        # 但层次化设计中，深度≥2的子模块内元件 PHYS_PAGE 是子模块内页码，不是主图页码。
        # 只有直接放置在顶层（路径中仅1个 SCH_1 层级）时，PHYS_PAGE 才是主图物理页码。
        phys_raw = attrs.get('PHYS_PAGE', '').strip()
        path_for_depth = page_path_raw or attrs.get('DRAWING', '')
        sch1_depth = len(re.findall(r'\(sch_1\)', path_for_depth, re.IGNORECASE))
        phys_page = f'PAGE{phys_raw}' if (phys_raw.isdigit() and sch1_depth <= 1) else ''
        components[refdes] = {
            'refdes':           refdes,
            'part_name':        part_name,
            'hq_code':          attrs.get('HQ_CODE', ''),
            'value':            attrs.get('VALUE', ''),
            'package':          attrs.get('PACKAGE', ''),
            'material':         attrs.get('MATERIAL', ''),
            'tolerance':        attrs.get('TOLERANCE', ''),
            'voltage':          attrs.get('VOLTAGE', ''),
            'current':          attrs.get('CURRENT', ''),
            'power':            attrs.get('POWER', ''),
            'bom_option':       attrs.get('BOM_OPTION', ''),
            'bom_cost':         attrs.get('BOM_COST', ''),
            'room':             attrs.get('ROOM', ''),
            'drawing':          attrs.get('DRAWING', ''),
            'page_path_raw':    page_path_raw,
            'page_path_source': page_path_source,
            'page':             phys_page or logical_page,
            'page_logical':     logical_page,
            'page_real':        phys_page,
            'comp_type':        _get_comp_type(refdes, part_name),
        }
    return components


def parse_pstxnet(content: str) -> Dict[str, List[dict]]:
    text = _join_continuations(content)
    nets = {}
    node_re     = re.compile(r'NODE_NAME\s+(\S+)\s+(\S+)')
    pin_name_re = re.compile(r"'([^']+)'\s*:")
    for block in re.split(r'\nNET_NAME\n', text)[1:]:
        m = re.search(r"'([^']+)'", block)
        if not m:
            continue
        net_name = m.group(1)
        nodes = []
        matches = list(node_re.finditer(block))
        for idx, nm in enumerate(matches):
            next_start = matches[idx + 1].start() if idx + 1 < len(matches) else len(block)
            after = block[nm.end():next_start]
            pn_match = pin_name_re.search(after)
            nodes.append({
                'refdes':   nm.group(1),
                'pin':      nm.group(2),
                'pin_name': pn_match.group(1) if pn_match else nm.group(2),
            })
        if nodes:
            nets[net_name] = nodes
    return nets


def parse_all(prt_content: str, net_content: str):
    components = parse_pstxprt(prt_content)
    nets       = parse_pstxnet(net_content)
    comp_nets: Dict[str, Dict[str, str]] = {}
    for net_name, nodes in nets.items():
        for node in nodes:
            rd = node['refdes']
            if rd not in comp_nets:
                comp_nets[rd] = {}
            comp_nets[rd][node['pin']] = net_name
    for refdes, comp in components.items():
        comp['nets'] = comp_nets.get(refdes, {})
    return components, nets, comp_nets


def _is_depop_option(bom_option: str) -> bool:
    return str(bom_option or '').strip().upper() in {'DEPOP', 'DNP'}


def _display_bom_option(bom_option: str) -> str:
    v = str(bom_option or '').strip().upper()
    return v or '默认'


# ══════════════════════════════════════════════════════════
# 二、BOM 分析
# ══════════════════════════════════════════════════════════

COMP_TYPE_CN = {
    'CAP': '电容', 'CAP_POL': '电解/钽电容', 'RES': '电阻',
    'IND': '电感/磁珠', 'IC': 'IC 芯片', 'CONN': '连接器',
    'DIODE': '二极管', 'LED': 'LED', 'FET': 'MOS/FET',
    'BJT': '三极管', 'XTAL': '晶振', 'FUSE': '保险丝',
    'SWITCH': '开关', 'TESTPOINT': '测试点', 'TRANSFORMER': '变压器',
}
_TYPE_ORDER = list(COMP_TYPE_CN.keys())


def build_bom(components: Dict):
    detail_normal, detail_depop = [], []
    for comp in components.values():
        ctype = comp.get('comp_type', '')
        row = {
            '位号':          comp['refdes'],
            '料号':          comp.get('hq_code', ''),
            '描述':          comp.get('part_name', ''),
            '值':            comp.get('value', ''),
            '封装':          comp.get('package', ''),
            '耐压/额定电压': comp.get('voltage', ''),
            '额定功率':      comp.get('power', ''),
            '精度':          comp.get('tolerance', ''),
            '材质':          comp.get('material', ''),
            '类型':          COMP_TYPE_CN.get(ctype, ctype),
            '_ctype':        ctype,
            '页面':          comp.get('page', ''),
            'ROOM':          comp.get('room', ''),
        }
        (detail_depop if _is_depop_option(comp.get('bom_option', '')) else detail_normal).append(row)

    def _merge(detail):
        if not detail:
            return []
        groups = {}
        for row in detail:
            key = row['料号'] or row['描述']
            if key not in groups:
                groups[key] = {
                    '料号': row['料号'], '位号列表': [], '数量': 0,
                    '描述': row['描述'], '值': row['值'], '封装': row['封装'],
                    '耐压': row['耐压/额定电压'], '额定功率': row['额定功率'],
                    '精度': row['精度'], '材质': row['材质'],
                    '类型': row['类型'], '_ctype': row['_ctype'],
                }
            groups[key]['位号列表'].append(row['位号'])
            groups[key]['数量'] += 1
        merged = list(groups.values())
        merged.sort(key=lambda r: (
            _TYPE_ORDER.index(r['_ctype']) if r['_ctype'] in _TYPE_ORDER else 99, r['料号']))
        for i, r in enumerate(merged, 1):
            r['序号'] = i
            r['位号列表'] = ', '.join(sorted(r['位号列表'], key=_natural_sort_key))
            del r['_ctype']
        return merged

    def _clean(rows):
        return [{k: v for k, v in r.items() if k != '_ctype'} for r in rows]

    return _clean(detail_normal), _clean(detail_depop), _merge(detail_normal), _merge(detail_depop)


# ══════════════════════════════════════════════════════════
# 三、网络分析
# ══════════════════════════════════════════════════════════

_DIFF_SUFFIX_PAIRS = [
    ('_P', '_N'), ('_DP', '_DN'), ('.P', '.N'),
    ('_TXPLUS', '_TXMINUS'), ('_RXPLUS', '_RXMINUS'),
]


def _collect_diff_pairs(nets: Dict) -> Dict[str, dict]:
    diff_pairs: Dict[str, dict] = {}
    upper_map = {name.upper(): name for name in nets}
    for net_name in nets:
        upper = net_name.upper()
        for pos_sfx, neg_sfx in _DIFF_SUFFIX_PAIRS:
            pu, nu = pos_sfx.upper(), neg_sfx.upper()
            if upper.endswith(pu):
                partner = upper_map.get(upper[:-len(pu)] + nu)
                if partner:
                    diff_pairs[net_name[:-len(pos_sfx)]] = {'P': net_name, 'N': partner}
                    break
            elif upper.endswith(nu):
                partner = upper_map.get(upper[:-len(nu)] + pu)
                base = net_name[:-len(neg_sfx)]
                if partner and base not in diff_pairs:
                    diff_pairs[base] = {'P': partner, 'N': net_name}
                    break
    return diff_pairs


def analyze_networks(nets: Dict, components: Dict) -> dict:
    single_node = {k: v for k, v in nets.items() if len(v) == 1}
    gnd_nets    = {k: v for k, v in nets.items() if _net_is_gnd(k)}
    power_nets  = {k: v for k, v in nets.items() if _net_is_power(k) and k not in gnd_nets}
    diff_pairs  = _collect_diff_pairs(nets)
    page_counter: Counter = Counter()
    for comp in components.values():
        page_counter[comp.get('page', '') or 'UNKNOWN'] += 1
    return {
        'total': len(nets), 'single_node': single_node,
        'gnd_nets': gnd_nets, 'power_nets': power_nets,
        'diff_pairs': diff_pairs, 'page_counter': page_counter,
    }


# ══════════════════════════════════════════════════════════
# 四、DRC 设计检查
# ══════════════════════════════════════════════════════════

_VALID_BOM_OPTIONS = {'', 'DEPOP', 'OPTION', 'MAIN_PLD', 'MAIN', 'ALT', 'DNP'}
_FUZZY_KEYWORDS    = sorted(opt for opt in _VALID_BOM_OPTIONS if opt)


def _edit_distance(a: str, b: str) -> int:
    a, b = a.upper(), b.upper()
    if a == b: return 0
    if not a:  return len(b)
    if not b:  return len(a)
    dp = list(range(len(b) + 1))
    for i, ca in enumerate(a):
        prev = dp[:]
        dp[0] = i + 1
        for j, cb in enumerate(b):
            dp[j+1] = min(prev[j] + (0 if ca == cb else 1), dp[j]+1, prev[j+1]+1)
    return dp[len(b)]


def check_drc(components: Dict, nets: Dict) -> dict:
    missing_hq, missing_val, missing_pkg, tbd_attrs, single_pin, unnamed = [], [], [], [], [], []
    bom_option_components = []
    for refdes, comp in components.items():
        ctype = comp.get('comp_type', '')
        if ctype == 'TESTPOINT':
            continue
        base = {'位号': refdes, '类型': COMP_TYPE_CN.get(ctype, ctype), '页面': comp.get('page', '')}
        if not comp.get('hq_code'):  missing_hq.append(base.copy())
        if not comp.get('value'):    missing_val.append(base.copy())
        if not comp.get('package'):  missing_pkg.append(base.copy())
        for attr in ('voltage', 'current', 'power'):
            val = comp.get(attr, '')
            if val and 'TBD' in val.upper():
                tbd_attrs.append({'位号': refdes, '属性': attr.upper(), '当前值': val,
                                  '类型': COMP_TYPE_CN.get(ctype, ctype), '页面': comp.get('page', '')})
        # BOM_OPTION 元件清单
        bom_option = str(comp.get('bom_option', '') or '').strip().upper()
        if bom_option:
            bom_option_components.append({
                '位号': refdes,
                '类型': COMP_TYPE_CN.get(ctype, ctype),
                'BOM_OPTION值': bom_option,
                '是否DEPOP': '是' if _is_depop_option(bom_option) else '否',
                '页面': comp.get('page', ''),
            })

    for net_name, nodes in nets.items():
        if len(nodes) == 1:
            n = nodes[0]
            comp = components.get(n['refdes'], {})
            if comp.get('comp_type') != 'TESTPOINT' and not re.search(r'^UNNAMED_', net_name, re.I):
                single_pin.append({'网络名': net_name, '连接元件': n['refdes'],
                                    '引脚': n['pin_name'],
                                    '页面': comp.get('page', '')})
        if re.search(r'^UNNAMED_', net_name, re.I):
            unnamed.append({'网络名': net_name, '节点数': len(nodes)})

    # 计算每个 BOM_OPTION 值的拼写风险
    risk_per_value: Dict[str, str] = {}
    for val in set(str(comp.get('bom_option', '') or '').strip().upper() for comp in components.values()):
        if not val:
            continue
        if val in _VALID_BOM_OPTIONS:
            risk_per_value[val] = '✅ 合法'
        else:
            min_d = min(_edit_distance(val, kw) for kw in _FUZZY_KEYWORDS)
            risk_per_value[val] = '❌ 疑似拼错' if min_d <= 2 else '⚠ 未知值'

    # 将风险信息写入每个元件行
    for item in bom_option_components:
        item['拼写风险'] = risk_per_value.get(item['BOM_OPTION值'], '')

    return {
        'missing_hq_code': missing_hq, 'missing_value': missing_val,
        'missing_package': missing_pkg, 'tbd_attrs': tbd_attrs,
        'single_pin_nets': single_pin, 'unnamed_nets': unnamed,
        'bom_option_components': sorted(bom_option_components, key=lambda r: _natural_sort_key(r['位号'])),
    }



# ══════════════════════════════════════════════════════════
# 五、电容降额分析
# ══════════════════════════════════════════════════════════

# PG/OD/OC 信号网络模式：这类网络电压由外部上拉决定，不推断
_OD_SKIP_PATTERNS = re.compile(
    r'\bPG\b|PGOOD|_PG_|_PGD\b|PG_N|PWRGD|POWER_GOOD'
    r'|\bFAULT\b|_FAULT|VR_FAULT'
    r'|\bALERT\b|_ALERT|SMBALERT'
    r'|\bSDA\b|\bSCL\b'
    r'|\bOC_N\b|_OC\b'
    r'|\bPRSNT\b|\bPRESENT\b'
    r'|\bINT_N\b|\bIRQ_N\b',
    re.IGNORECASE,
)


def _split_net_tokens(net_name: str) -> List[str]:
    return [tok for tok in re.split(r'[_./-]+', (net_name or '').upper()) if tok]


def _first_net_token(net_name: str) -> str:
    tokens = _split_net_tokens(net_name)
    return tokens[0] if tokens else (net_name or '').upper()


_POWER_TOKEN_RE = re.compile(
    r'(?:VCC|VDD|VBAT|VCORE|VCCIO|PVDD|PVCC|AVDD|DVDD|VBUS)[A-Z0-9]*',
    re.IGNORECASE,
)
_GROUND_TOKEN_RE = re.compile(
    r'(?:GND|AGND|SGND|PGND|DGND|VSS[A-Z0-9]*)',
    re.IGNORECASE,
)


def _token_is_power(token: str) -> bool:
    m = re.fullmatch(r'P?(\d+)V(\d*)', token.upper())
    if m:
        return True
    return bool(_POWER_TOKEN_RE.fullmatch(token))


def _token_is_ground(token: str) -> bool:
    return bool(_GROUND_TOKEN_RE.fullmatch(token))


def _net_is_power(net: str) -> bool:
    return _token_is_power(_first_net_token(net))


def _net_is_gnd(net: str) -> bool:
    return _token_is_ground(_first_net_token(net))


def _parse_voltage_from_token(token: str) -> Optional[float]:
    m = re.fullmatch(r'P?(\d+)V(\d*)', token.upper())
    if not m:
        return None
    int_part, frac_part = m.groups()
    return float(f'{int_part}.{frac_part}') if frac_part else float(int_part)


def _infer_voltage(net_name: str) -> Optional[float]:
    """从网络名首 token 推断电压（新版：基于 token 解析，不误判 PG/OD 信号）"""
    if _OD_SKIP_PATTERNS.search(net_name):
        return None
    token = _first_net_token(net_name)
    if _token_is_ground(token):
        return 0.0
    return _parse_voltage_from_token(token)


def _is_od_net(net_name: str) -> bool:
    return bool(_OD_SKIP_PATTERNS.search(net_name))


def _matches_prefix_with_boundary(name: str, prefix: str) -> bool:
    if not prefix:
        return False
    name = (name or '').upper()
    prefix = prefix.upper()
    if not name.startswith(prefix):
        return False
    return len(name) == len(prefix) or name[len(prefix)] in '_./-'


def _match_custom_voltage(net_name: str, custom_volt_map: Optional[Dict]) -> Optional[float]:
    if not custom_volt_map:
        return None
    best: Optional[Tuple[int, float]] = None
    for key, volt in custom_volt_map.items():
        prefix = str(key).strip().upper()
        if prefix and _matches_prefix_with_boundary(net_name, prefix):
            if best is None or len(prefix) > best[0]:
                best = (len(prefix), float(volt))
    return best[1] if best else None


def _collect_component_nets(nets: Dict) -> Dict[str, List[str]]:
    comp_nets: Dict[str, List[str]] = defaultdict(list)
    for net_name, nodes in nets.items():
        for node in nodes:
            comp_nets[node['refdes']].append(net_name)
    return comp_nets


def _unique_component_nets(comp_nets: Dict, refdes: str) -> List[str]:
    return list(dict.fromkeys(comp_nets.get(refdes, [])))


def _find_ac_coupling_candidates(components: Dict,
                                  comp_nets: Dict[str, List[str]],
                                  nets: Dict) -> Dict[str, dict]:
    """查找 AC 耦合电容候选：两端都接差分对同极性 net，且有镜像电容"""
    upper_map = {name.upper(): name for name in nets}

    def _get_diff_info(net_name):
        upper = net_name.upper()
        for pos_sfx, neg_sfx in _DIFF_SUFFIX_PAIRS:
            pu, nu = pos_sfx.upper(), neg_sfx.upper()
            if upper.endswith(pu):
                partner = upper_map.get(upper[:-len(pu)] + nu)
                if partner:
                    return {'polarity': 'P', 'partner': partner}
            elif upper.endswith(nu):
                partner = upper_map.get(upper[:-len(nu)] + pu)
                if partner:
                    return {'polarity': 'N', 'partner': partner}
        return None

    cap_pairs: Dict[str, Tuple[str, str]] = {}
    caps_by_pair: Dict[frozenset, List[str]] = defaultdict(list)
    for refdes, comp in components.items():
        if comp.get('comp_type') not in ('CAP', 'CAP_POL'):
            continue
        unique_nets = _unique_component_nets(comp_nets, refdes)
        if len(unique_nets) != 2:
            continue
        na, nb = unique_nets
        if _net_is_power(na) or _net_is_power(nb) or _net_is_gnd(na) or _net_is_gnd(nb):
            continue
        cap_pairs[refdes] = (na, nb)
        caps_by_pair[frozenset((na, nb))].append(refdes)

    candidates: Dict[str, dict] = {}
    for refdes, (na, nb) in cap_pairs.items():
        ia = _get_diff_info(na)
        ib = _get_diff_info(nb)
        if not ia or not ib or ia['polarity'] != ib['polarity']:
            continue
        partner_pair = frozenset((ia['partner'], ib['partner']))
        mirror_caps = sorted([c for c in caps_by_pair.get(partner_pair, []) if c != refdes],
                             key=_natural_sort_key)
        if not mirror_caps:
            continue
        candidates[refdes] = {
            'nets': (na, nb), 'mirror_nets': sorted(partner_pair, key=_natural_sort_key),
            'mirror_caps': mirror_caps, 'polarity': ia['polarity'],
        }
    return candidates


def _calc_board_max_voltage(nets: Dict, custom_volt_map: Optional[Dict]) -> float:
    """扫描全板所有网络名，推断板级最高工作电压（用于快速 pass 高额定电容）"""
    max_v = 0.0
    for net_name in nets:
        v = _match_custom_voltage(net_name, custom_volt_map)
        if v is None:
            v = _infer_voltage(net_name)
        if v is not None and v > max_v:
            max_v = v
    return max_v


def analyze_derating(components: Dict, nets: Dict,
                     pct: float = 70.0,
                     custom_volt_map: Optional[Dict[str, float]] = None,
                     include_depop: bool = False) -> List[dict]:
    """pct: 工作电压上限占额定电压的百分比（默认 70%）"""
    comp_nets = _collect_component_nets(nets)
    ac_coupling_caps = _find_ac_coupling_candidates(components, comp_nets, nets)
    board_max_v = _calc_board_max_voltage(nets, custom_volt_map)

    rows = []
    for refdes, comp in components.items():
        ctype = comp.get('comp_type', '')
        if ctype not in ('CAP', 'CAP_POL'):
            continue
        if not include_depop and _is_depop_option(comp.get('bom_option', '')):
            continue
        connected_nets = _unique_component_nets(comp_nets, refdes)
        rated_str = comp.get('voltage', '')
        source_type = ''
        max_v, from_net, derating = None, '', None

        if not rated_str:
            status = '⚪ 无额定电压'
        else:
            m = re.match(r'([\d.]+)\s*V', rated_str.strip(), re.I)
            rated_v = float(m.group(1)) if m else None
            if rated_v is None:
                status = '⚪ 无法解析额定电压'
            elif board_max_v > 0 and rated_v * (pct / 100) >= board_max_v:
                # 额定电压 × 降额比 ≥ 板级最高电压 → 无论接哪个网络都安全，直接 pass
                threshold_v = rated_v * (pct / 100)
                status = (f'✅ 板级直通 (额定{rated_v:.0f}V×{pct:.0f}%={threshold_v:.1f}V'
                          f' ≥ 板级最高{board_max_v:.1f}V)')
                source_type = '板级直通'
            elif refdes in ac_coupling_caps:
                status = '⚪ 疑似 AC 耦合电容，不推断电压'
                source_type = 'AC 耦合候选'
            else:
                known_nets = []
                ground_present = False
                for net_name in connected_nets:
                    if _net_is_gnd(net_name):
                        ground_present = True
                    # PG/OD 信号：标记为特殊，跳过电压推断
                    if _is_od_net(net_name) and not _net_is_gnd(net_name):
                        continue
                    v = _match_custom_voltage(net_name, custom_volt_map)
                    src = 'custom_map' if v is not None else ''
                    if v is None:
                        v = _infer_voltage(net_name)
                        if v is not None:
                            src = 'net_token'
                    if v is None:
                        continue
                    if v == 0:
                        ground_present = True
                    known_nets.append((net_name, float(v), src))

                positives: Dict[float, Tuple[str, str]] = {}
                for net_name, v, src in known_nets:
                    if v > 0:
                        positives.setdefault(round(v, 6), (net_name, src))

                od_nets = [n for n in connected_nets if _is_od_net(n) and not _net_is_gnd(n)]

                if not ground_present:
                    if od_nets:
                        status = f'⚪ PG/OD信号（{od_nets[0]}），工作电压由上拉决定，请手动确认'
                    else:
                        status = '⚪ 无法判断（未连接地）'
                elif not positives:
                    if od_nets:
                        status = f'⚪ PG/OD信号（{od_nets[0]}），工作电压由上拉决定，请手动确认'
                    else:
                        status = '⚪ 无法推断工作电压'
                elif len(positives) > 1:
                    status = '⚪ 无法判断（连接多个不同电位）'
                else:
                    rounded_v, (from_net, src) = next(iter(positives.items()))
                    max_v = rounded_v
                    source_type = '自定义映射' if src == 'custom_map' else '网络首 token'
                    usage_pct = max_v / rated_v * 100
                    derating = rated_v / max_v
                    if usage_pct <= pct:
                        status = f'✅ 合格 ({usage_pct:.0f}% ≤ {pct:.0f}%)'
                    else:
                        status = f'❌ 不合格 ({usage_pct:.0f}% > {pct:.0f}%)'

        rows.append({
            '位号':            refdes,
            '值':              comp.get('value', ''),
            '封装':            comp.get('package', ''),
            '类型':            COMP_TYPE_CN.get(ctype, ctype),
            '额定电压':        rated_str,
            '推断工作电压(V)': str(max_v) if max_v is not None else '',
            '推断来源网络':    from_net,
            '推断来源类型':    source_type,
            '所有连接网络':    ', '.join(connected_nets),
            '降额比':          f'{derating:.2f}' if derating is not None else '',
            '状态':            status,
            '页面':            comp.get('page', ''),
            'DEPOP':           'Y' if _is_depop_option(comp.get('bom_option', '')) else '',
        })
    rows.sort(key=lambda r: (
        0 if r['状态'].startswith('❌') else 1 if r['状态'].startswith('✅') else 2,
        _natural_sort_key(r.get('位号', '')),
    ))
    return rows



# ══════════════════════════════════════════════════════════
# 六、电阻检查（上拉 / 下拉 / 串阻 / OD/OC / 芯片Pin总览）
# ══════════════════════════════════════════════════════════

def _parse_ohms(value_str: str) -> Optional[float]:
    if not value_str:
        return None
    s = re.sub(r'\s', '', value_str.upper())
    s = s.replace('Ω', 'R').replace('OHM', 'R').replace('OHMS', 'R')
    m = re.match(r'^([\d.]+)([KMGR]?)$', s)
    if m:
        val = float(m.group(1))
        return val * {'K': 1e3, 'M': 1e6, 'G': 1e9, 'R': 1, '': 1}.get(m.group(2), 1)
    # 支持 4K7 → 4.7k 写法
    embedded = re.match(r'^(\d+)([KMGR])(\d+)$', s)
    if embedded:
        val = float(f'{embedded.group(1)}.{embedded.group(3)}')
        return val * {'K': 1e3, 'M': 1e6, 'G': 1e9, 'R': 1}.get(embedded.group(2), 1)
    return None


_CHIP_REFDES_RE = re.compile(r'^(?:XU|PU|U)[A-Z0-9]+$', re.IGNORECASE)


def _is_chip_component(refdes: str, comp: Dict) -> bool:
    return comp.get('comp_type') == 'IC' and bool(_CHIP_REFDES_RE.match(refdes or ''))


# OD/OC 信号名关键词（用于多证据判定）
_OD_STRONG_TOKENS = {'SDA', 'SCL', 'SMBALERT', 'SMBDAT', 'SMBDATA', 'SMBCLK', 'OD', 'OC'}
_OD_WEAK_TOKENS = {'ALERT', 'FAULT', 'IRQ', 'INT', 'PGOOD', 'PWROK', 'PWRGD', 'PRSNT', 'PRESENT'}


def _od_oc_evidence_from_name(value: str, source_label: str) -> List[Tuple[str, str]]:
    tokens = set(re.findall(r'[A-Z0-9]+', (value or '').upper()))
    evidence = []
    for tok in _OD_STRONG_TOKENS:
        if tok in tokens:
            evidence.append(('strong', f'{source_label} 含 {tok}'))
    for tok in _OD_WEAK_TOKENS:
        if tok in tokens:
            evidence.append(('weak', f'{source_label} 含 {tok}'))
    return evidence


def _classify_od_oc_evidence(net_name: str, nodes: List[dict],
                              components: Dict) -> Optional[Dict]:
    evidence = []
    chip_nodes = []
    for node in nodes:
        refdes = node.get('refdes', '')
        comp = components.get(refdes, {})
        if not _is_chip_component(refdes, comp):
            continue
        chip_nodes.append(node)
        evidence.extend(_od_oc_evidence_from_name(
            node.get('pin_name', node.get('pin', '')), f'{refdes}.{node.get("pin", "")}'))
    if not chip_nodes:
        return None
    evidence.extend(_od_oc_evidence_from_name(net_name, '网络名'))
    strong = [t for lvl, t in evidence if lvl == 'strong']
    weak = [t for lvl, t in evidence if lvl == 'weak']
    if not strong and len(weak) < 2:
        return None
    unique_evidence = list(dict.fromkeys(strong + weak))
    chip_pins = ', '.join(dict.fromkeys(
        f'{n["refdes"]}.{n["pin"]}({n.get("pin_name", n["pin"])})' for n in chip_nodes))
    return {
        '芯片引脚': chip_pins,
        '判定依据': '; '.join(unique_evidence[:6]),
        'confidence': 'medium' if strong else 'low',
    }


def _classify_series_bias_ratio(series_ohms, bias_ohms):
    if series_ohms is None or bias_ohms is None or bias_ohms <= 0:
        return None, '⚪ 阻值缺失，无法计算'
    ratio = series_ohms / bias_ohms
    if bias_ohms < 1000 and ratio > 0.1:
        return ratio, '❌ 高风险'
    if ratio >= 0.33:
        return ratio, '❌ 高风险'
    if ratio > 0.1:
        return ratio, '⚠️ 关注'
    return ratio, '✅ 正常'


def _format_entry_list(entries: List[dict], key: str) -> str:
    return ', '.join(dict.fromkeys(str(e.get(key, '')) for e in entries if e.get(key, '') != ''))


def analyze_resistors(components: Dict, nets: Dict, exclude_depop: bool = True) -> dict:
    """检测上拉/下拉/串阻相关设计问题，含双向扫描和芯片Pin总览"""
    pullups:   Dict[str, list] = defaultdict(list)
    pulldowns: Dict[str, list] = defaultdict(list)
    series_list: list = []
    series_by_net: Dict[str, list] = defaultdict(list)
    indirect_pullups: Dict[str, list] = defaultdict(list)
    indirect_pulldowns: Dict[str, list] = defaultdict(list)
    node_lookup: Dict[Tuple[str, str], str] = {}

    for net_name, nodes in nets.items():
        for node in nodes:
            node_lookup[(node['refdes'], node['pin'])] = node.get('pin_name', node['pin'])

    for refdes, comp in components.items():
        if comp.get('comp_type') != 'RES':
            continue
        if exclude_depop and _is_depop_option(comp.get('bom_option', '')):
            continue
        pin_nets = list(dict.fromkeys(comp.get('nets', {}).values()))
        if len(pin_nets) != 2:
            continue
        net_a, net_b = pin_nets[0], pin_nets[1]
        ohms = _parse_ohms(comp.get('value', ''))
        val_str = comp.get('value', '')
        page = comp.get('page', '')
        bom_option = comp.get('bom_option', '')

        a_pwr, b_pwr = _net_is_power(net_a), _net_is_power(net_b)
        a_gnd, b_gnd = _net_is_gnd(net_a),   _net_is_gnd(net_b)

        entry_base = {'refdes': refdes, 'ohms': ohms, 'value': val_str, 'page': page, 'bom_option': bom_option}
        if a_pwr and not b_pwr and not b_gnd:
            pullups[net_b].append({**entry_base, 'power_net': net_a})
        elif b_pwr and not a_pwr and not a_gnd:
            pullups[net_a].append({**entry_base, 'power_net': net_b})
        elif a_gnd and not b_gnd and not b_pwr:
            pulldowns[net_b].append(entry_base.copy())
        elif b_gnd and not a_gnd and not a_pwr:
            pulldowns[net_a].append(entry_base.copy())
        elif not a_pwr and not b_pwr and not a_gnd and not b_gnd:
            series_list.append({**entry_base, 'net_a': net_a, 'net_b': net_b})
            series_by_net[net_a].append({**entry_base, 'other_net': net_b})
            series_by_net[net_b].append({**entry_base, 'other_net': net_a})

    # ── 检查1：重复上拉 ─────────────────────────────────
    dup_pullups = []
    for sig_net, pu_list in sorted(pullups.items()):
        if len(pu_list) < 2:
            continue
        group = sorted(pu_list, key=lambda r: _natural_sort_key(r.get('refdes', '')))
        dup_pullups.append({
            '信号网络':  sig_net,
            '上拉数量':  len(group),
            '位号':      ', '.join(r['refdes'] for r in group),
            '阻值':      ', '.join(r['value']  for r in group),
            '上拉电源':  ', '.join(dict.fromkeys(r['power_net'] for r in group)),
            'BOM_OPTION': ', '.join(dict.fromkeys(_display_bom_option(r['bom_option']) for r in group)),
            '页面':      ', '.join(dict.fromkeys(r['page'] for r in group)),
        })

    # ── 检查2：重复下拉 ─────────────────────────────────
    dup_pulldowns = []
    for sig_net, pd_list in sorted(pulldowns.items()):
        if len(pd_list) < 2:
            continue
        group = sorted(pd_list, key=lambda r: _natural_sort_key(r.get('refdes', '')))
        dup_pulldowns.append({
            '信号网络': sig_net,
            '下拉数量': len(group),
            '位号':     ', '.join(r['refdes'] for r in group),
            '阻值':     ', '.join(r['value']  for r in group),
            'BOM_OPTION': ', '.join(dict.fromkeys(_display_bom_option(r['bom_option']) for r in group)),
            '页面':     ', '.join(dict.fromkeys(r['page'] for r in group)),
        })

    # ── 检查3：串阻 + 偏置电阻分压风险（双向扫描）─────
    divider_risks = []
    seen_pairs: set = set()
    seen_indirect: set = set()
    for sr in sorted(series_list, key=lambda r: _natural_sort_key(r.get('refdes', ''))):
        for bias_net, affected_net in ((sr['net_a'], sr['net_b']), (sr['net_b'], sr['net_a'])):
            for bias_kind, bias_map, indirect_map in [
                ('上拉', pullups, indirect_pullups),
                ('下拉', pulldowns, indirect_pulldowns),
            ]:
                for bias in bias_map.get(bias_net, []):
                    # 记录间接偏置
                    ik = (affected_net, bias_kind, bias['refdes'], sr['refdes'])
                    if ik not in seen_indirect:
                        seen_indirect.add(ik)
                        indirect_map[affected_net].append({
                            **bias, 'via_refdes': sr['refdes'],
                            'via_value': sr['value'], 'via_ohms': sr['ohms'],
                            'source_net': bias_net, 'other_net': affected_net,
                        })
                    pair_key = (sr['refdes'], bias['refdes'], bias_kind, bias_net, affected_net)
                    if pair_key in seen_pairs:
                        continue
                    seen_pairs.add(pair_key)
                    ratio, status = _classify_series_bias_ratio(sr['ohms'], bias.get('ohms'))
                    ref_net = bias.get('power_net', '') if bias_kind == '上拉' else 'GND'
                    pages = ', '.join(dict.fromkeys(v for v in [sr.get('page', ''), bias.get('page', '')] if v))
                    divider_risks.append({
                        '串阻位号':    sr['refdes'],
                        '串阻值':      sr['value'],
                        '串阻网络A':   sr['net_a'],
                        '串阻网络B':   sr['net_b'],
                        '偏置类型':    bias_kind,
                        '偏置位号':    bias['refdes'],
                        '偏置值':      bias['value'],
                        '偏置所在网络': bias_net,
                        '偏置参考网络': ref_net,
                        '受影响网络':  affected_net,
                        '串/偏置比':   f'{ratio:.3f}' if ratio is not None else '',
                        '偏置 < 1k':  '是' if (bias.get('ohms') or 0) < 1000 else '否',
                        '说明':        f'{bias_kind}位于 {bias_net} 侧，通过 {sr["refdes"]} 影响 {affected_net}',
                        '状态':        status,
                        '页面':        pages,
                    })
    divider_risks.sort(key=lambda r: (
        0 if r['状态'].startswith('❌') else 1 if r['状态'].startswith('⚠') else 2,
        _natural_sort_key(r.get('串阻位号', '')),
    ))

    # ── 检查4：OD/OC 信号缺上拉（多证据判定）──────────
    od_missing = []
    for net_name in sorted(nets.keys()):
        if _net_is_power(net_name) or _net_is_gnd(net_name):
            continue
        nodes = nets[net_name]
        evidence = _classify_od_oc_evidence(net_name, nodes, components)
        if not evidence:
            continue
        if pullups.get(net_name) or indirect_pullups.get(net_name):
            continue
        od_missing.append({
            '网络名':   net_name,
            '节点数':   len(nodes),
            '连接元件': ', '.join(dict.fromkeys(n['refdes'] for n in nodes[:6])),
            '芯片引脚': evidence['芯片引脚'],
            '判定依据': evidence['判定依据'],
            '上拉状态': '未找到直接上拉/隔串阻上拉',
            '说明':     '疑似 OD/OC 信号，未找到上拉电阻',
        })

    # ── 芯片 Pin 电阻状态总览 ────────────────────────────
    chip_pin_rows = []
    for refdes, comp in sorted(components.items(), key=lambda item: _natural_sort_key(item[0])):
        if not _is_chip_component(refdes, comp):
            continue
        for pin, net_name in sorted(comp.get('nets', {}).items(), key=lambda item: _natural_sort_key(item[0])):
            pin_name = node_lookup.get((refdes, pin), pin)
            s_entries = series_by_net.get(net_name, [])
            pu_entries = pullups.get(net_name, [])
            pd_entries = pulldowns.get(net_name, [])
            ipu_entries = indirect_pullups.get(net_name, [])
            ipd_entries = indirect_pulldowns.get(net_name, [])
            chip_pin_rows.append({
                '芯片位号': refdes,
                '引脚':     pin,
                '引脚名':   pin_name,
                '网络名':   net_name,
                '有串阻':   '是' if s_entries else '否',
                '串阻数量': len(s_entries),
                '串阻位号': _format_entry_list(s_entries, 'refdes'),
                '串阻另一端': _format_entry_list(s_entries, 'other_net'),
                '有上拉':   '是' if pu_entries else '否',
                '上拉数量': len(pu_entries),
                '上拉位号': _format_entry_list(pu_entries, 'refdes'),
                '上拉电源': _format_entry_list(pu_entries, 'power_net'),
                '隔串阻上拉数量': len(ipu_entries),
                '隔串阻上拉位号': _format_entry_list(ipu_entries, 'refdes'),
                '有下拉':   '是' if pd_entries else '否',
                '下拉数量': len(pd_entries),
                '下拉位号': _format_entry_list(pd_entries, 'refdes'),
                '隔串阻下拉数量': len(ipd_entries),
                '页面':     comp.get('page', ''),
            })

    return {
        'dup_pullups':         dup_pullups,
        'dup_pulldowns':       dup_pulldowns,
        'divider_risks':       divider_risks,
        'od_missing':          od_missing,
        'chip_pin_rows':       chip_pin_rows,
        'pullups':             dict(pullups),
        'pulldowns':           dict(pulldowns),
        'indirect_pullups':    dict(indirect_pullups),
        'indirect_pulldowns':  dict(indirect_pulldowns),
        'series_by_net':       dict(series_by_net),
    }



# ══════════════════════════════════════════════════════════
# 八、Excel 导出
# ══════════════════════════════════════════════════════════

_BL = PatternFill("solid", fgColor="1F4E79")
_OR = PatternFill("solid", fgColor="C55A11")
_GR = PatternFill("solid", fgColor="375623")
_GY = PatternFill("solid", fgColor="595959")
_RF = PatternFill("solid", fgColor="FFCCCC")
_WF = Font(color="FFFFFF", bold=True, size=10)
_BF = Font(bold=True, size=10)
_NF = Font(size=10)
_CA = Alignment(horizontal='center', vertical='center', wrap_text=True)
_LA = Alignment(horizontal='left',   vertical='center', wrap_text=True)
_TH = Side(style='thin')
_BD = Border(left=_TH, right=_TH, top=_TH, bottom=_TH)


def _xl_hdr(ws, row_idx, fill):
    for cell in ws[row_idx]:
        if cell.value is not None:
            cell.fill = fill; cell.font = _WF; cell.alignment = _CA; cell.border = _BD


def _xl_autowidth(ws, mx=50):
    for col in ws.columns:
        vals = [str(c.value or '') for c in col]
        ws.column_dimensions[col[0].column_letter].width = min(
            max((len(v) for v in vals), default=8) + 2, mx)


def _xl_write_rows(ws, rows: List[dict], fill, hl_col=None, freeze=True):
    if not rows:
        ws.append(['（无数据）']); return
    hdrs = list(rows[0].keys())
    ws.append(hdrs); _xl_hdr(ws, ws.max_row, fill)
    hl_idx = hdrs.index(hl_col) if hl_col in hdrs else None
    for row in rows:
        ws.append(list(row.values()))
        ri  = ws.max_row
        red = hl_idx is not None and '❌' in str(ws.cell(ri, hl_idx+1).value or '')
        for cell in ws[ri]:
            cell.border = _BD; cell.alignment = _LA; cell.font = _NF
            if red: cell.fill = _RF
    _xl_autowidth(ws)
    if freeze: ws.freeze_panes = 'A2'


def _xl_section(ws, title, fill):
    ws.append([title])
    for cell in ws[ws.max_row]:
        cell.fill = fill; cell.font = _WF; cell.border = _BD
    ws.append([])


def export_to_excel(data: dict, out_path: str) -> str:
    base, ext = os.path.splitext(out_path)
    n, path = 1, out_path
    while os.path.exists(path):
        path = f'{base}({n}){ext}'; n += 1

    wb = Workbook(); wb.remove(wb.active)
    project = data.get('project_name', '')
    na  = data.get('net_analysis', {})
    drc = data.get('drc', {})
    drt = data.get('derating', [])
    mn  = data.get('bom_normal_merged', [])
    md  = data.get('bom_depop_merged', [])
    res = data.get('resistor_analysis', {})

    # 概览
    ws = wb.create_sheet('概览')
    ws.column_dimensions['A'].width = 30; ws.column_dimensions['B'].width = 16
    drc_total = sum(len(v) for v in drc.values() if isinstance(v, list))
    fail = sum(1 for r in drt if r.get('状态', '').startswith('❌'))
    for label, val in [
        ('项目名称', project),
        ('贴装元件种类数', len(mn)),
        ('贴装元件总数',  sum(r.get('数量', 0) for r in mn)),
        ('DEPOP 元件种类数', len(md)),
        ('DEPOP 元件总数',  sum(r.get('数量', 0) for r in md)),
        ('网络总数', na.get('total', '')),
        ('单端网络数（疑似漏连）', len(na.get('single_node', {}))),
        ('电源网络数', len(na.get('power_nets', {}))),
        ('差分对数', len(na.get('diff_pairs', {}))),
        ('DRC 问题总数', drc_total),
        ('电容降额不合格数', fail),
    ]:
        ws.append([label, val])
    for row in ws.iter_rows():
        for cell in row:
            cell.border = _BD
            cell.font = _BF if cell.column == 1 else _NF
            cell.alignment = _LA

    # BOM
    ws = wb.create_sheet('BOM_贴装'); _xl_write_rows(ws, mn, _BL)
    ws = wb.create_sheet('BOM_DEPOP'); _xl_write_rows(ws, md, _OR)
    ws = wb.create_sheet('BOM_明细')
    all_d = [{'DEPOP': '', **r} for r in data.get('bom_normal_detail', [])] + \
            [{'DEPOP': 'Y', **r} for r in data.get('bom_depop_detail', [])]
    _xl_write_rows(ws, all_d, _GY)

    # 网络分析
    ws = wb.create_sheet('网络分析'); ws.freeze_panes = None
    _xl_section(ws, '电源网络', _BL)
    _xl_write_rows(ws, [{'网络名': k, '节点数': len(v)}
                        for k, v in sorted(na.get('power_nets', {}).items(), key=lambda x: -len(x[1]))],
                   _BL, freeze=False)
    ws.append([])
    _xl_section(ws, 'GND 网络', _GR)
    _xl_write_rows(ws, [{'网络名': k, '节点数': len(v)}
                        for k, v in sorted(na.get('gnd_nets', {}).items(), key=lambda x: -len(x[1]))],
                   _GR, freeze=False)
    ws.append([])
    _xl_section(ws, '差分对', _OR)
    _xl_write_rows(ws, [{'基础名': b, 'P端网络': pr['P'], 'N端网络': pr['N']}
                        for b, pr in sorted(na.get('diff_pairs', {}).items())],
                   _OR, freeze=False)
    ws.append([])
    _xl_section(ws, '单端网络（疑似漏连）', _GY)
    _xl_write_rows(ws, [{'网络名': k, '连接元件': v[0]['refdes'], '引脚': v[0]['pin_name']}
                        for k, v in sorted(na.get('single_node', {}).items())],
                   _GY, freeze=False)
    ws.append([])
    _xl_section(ws, '各页面元件数', _BL)
    _xl_write_rows(ws, [{'页面': p, '元件数': c}
                        for p, c in sorted(na.get('page_counter', {}).items())],
                   _BL, freeze=False)
    _xl_autowidth(ws)

    # 设计检查
    ws = wb.create_sheet('设计检查'); ws.freeze_panes = None
    for title, key, fill in [
        ('TBD 待确认属性', 'tbd_attrs',       _OR),
        ('缺少料号',       'missing_hq_code',  _RF),
        ('缺少 VALUE',     'missing_value',     _RF),
        ('缺少封装',       'missing_package',   _RF),
        ('单端网络',       'single_pin_nets',   _GY),
        ('未命名网络',     'unnamed_nets',      _GY),
        ('BOM_OPTION 元件清单（含拼写风险）', 'bom_option_components', _BL),
    ]:
        _xl_section(ws, title, fill)
        _xl_write_rows(ws, drc.get(key, []), fill, freeze=False)
        ws.append([])
    _xl_autowidth(ws)

    # 降额
    ws = wb.create_sheet('降额分析')
    _xl_write_rows(ws, drt, _BL, hl_col='状态')

    # 电阻检查
    ws = wb.create_sheet('电阻检查'); ws.freeze_panes = None
    for title, key, hl, fill in [
        ('串阻分压风险', 'divider_risks', '状态', _OR),
        ('重复上拉', 'dup_pullups', None, _BL),
        ('重复下拉', 'dup_pulldowns', None, _GY),
        ('OD/OC 缺上拉', 'od_missing', None, _GR),
    ]:
        _xl_section(ws, title, fill)
        _xl_write_rows(ws, res.get(key, []), fill, hl_col=hl, freeze=False)
        ws.append([])
    _xl_autowidth(ws)

    # 芯片 Pin 总览
    chip_rows = res.get('chip_pin_rows', [])
    if chip_rows:
        ws = wb.create_sheet('芯片Pin总览')
        _xl_write_rows(ws, chip_rows, _BL)

    wb.save(path)
    return path

